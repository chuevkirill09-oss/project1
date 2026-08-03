"""
Расчёт объёма пиломатериалов и вывод таблицы в Excel.
Данные вводит пользователь вручную через консоль.

Заголовки таблицы:
- Наименование / сорт
- Размеры т/ш/д (толщина, ширина, длина в мм)
- Количество шт
- Объем м3
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def input_items():
    """Спрашивает у пользователя данные по каждой позиции пиломатериала."""
    items = []
    print("Ввод пиломатериалов.")
    print("Чтобы закончить ввод — просто нажмите Enter вместо названия.\n")

    while True:
        name = input("Наименование / сорт: ").strip()
        if name == "":
            break

        while True:
            size_str = input("Размеры т/ш/д, мм (например 27/140/3000): ").strip()
            parts = size_str.split("/")
            if len(parts) == 3 and all(p.replace(".", "", 1).isdigit() for p in parts):
                break
            print("  Неверный формат! Введите как толщина/ширина/длина, например: 27/140/3000")

        while True:
            qty_str = input("Количество шт: ").strip()
            if qty_str.isdigit() and int(qty_str) > 0:
                qty = int(qty_str)
                break
            print("  Введите целое положительное число.")

        items.append((name, size_str, qty))
        print(f"  Добавлено: {name} | {size_str} | {qty} шт\n")

    return items

def calc_volume(size_str: str, qty: int) -> float:
    """Считает объём в м3 по размеру 'т/ш/д' (мм) и количеству штук."""
    t, w, l = (float(x) for x in size_str.split("/"))
    volume_one = (t / 1000) * (w / 1000) * (l / 1000)  # м3 одной штуки
    return round(volume_one * qty, 4)

def build_table(items):
    rows = []
    total = 0
    for name, size_str, qty in items:
        volume = calc_volume(size_str, qty)
        total += volume
        rows.append((name, size_str, qty, volume))
    return rows, round(total, 4)

def save_to_excel(rows, total, path="pilomaterialy.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Пиломатериалы"

    headers = ["Наименование / сорт", "Размеры т/ш/д", "Количество шт", "Объем м3"]
    ws.append(headers)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = border

    for r in rows:
        ws.append(r)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = center

    # строка "Итого"
    ws.append(["Итого", "", "", total])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.border = border
        cell.alignment = center

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    wb.save(path)
    print(f"\nФайл сохранён: {path}")

if __name__ == "__main__":
    items = input_items()

    if not items:
        print("Нет данных для сохранения.")
    else:
        rows, total = build_table(items)
        print()
        print(f"{'Наименование':30}{'Размеры т/ш/д':18}{'Кол-во шт':12}{'Объем м3':10}")
        for name, size, qty, vol in rows:
            print(f"{name:30}{size:18}{qty:<12}{vol}")
        print(f"\nИтого объём: {total} м3")

        save_to_excel(rows, total)